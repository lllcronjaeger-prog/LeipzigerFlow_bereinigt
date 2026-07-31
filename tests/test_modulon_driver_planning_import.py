from datetime import date

from openpyxl import Workbook

from leipzigerflow.imports.modulon_resource_planner import build_preview


def test_modulon_preview_reads_daily_statuses(tmp_path):
    path = tmp_path / "Ressourcenplaner.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Zusatzinformationen"
    ws["C2"] = "07.2026"
    headers = ["Fahrernr.", "Pers.-Nr.", "Vorname", "Nachname", "Pers.-Nr.", "Gruppe", "Niederlassung", "# Tage"]
    for col, value in enumerate(headers, 1):
        ws.cell(6, col, value)
    ws.cell(6, 9, date(2026, 7, 1))
    ws.cell(6, 10, date(2026, 7, 2))
    ws.cell(8, 1, "0048")
    ws.cell(8, 3, "Michael")
    ws.cell(8, 4, "Rose")
    ws.cell(8, 6, "Nahverkehr")
    ws.cell(8, 7, "Ettlingen")
    ws.cell(8, 9, "FW")
    ws.cell(8, 10, "AF")
    wb.create_sheet("Feiertage im Zeitraum")
    wb.save(path)

    preview = build_preview(path)

    assert preview.month == date(2026, 7, 1)
    assert len(preview.rows) == 2
    assert preview.rows[0].mapped_status == "Freiwoche"
    assert preview.rows[1].mapped_status == "Modulon: AF"
    assert preview.unknown_statuses == {"AF"}


def test_modulon_matches_middle_name_and_polish_characters():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401 - registriert alle Tabellen
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningRow

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as session:
        celny = Driver(first_name="Maciej Jan", last_name="Celny", city="Leipzig")
        gawronski = Driver(first_name="Paweł", last_name="Gawroński", city="Leipzig")
        session.add_all([celny, gawronski])
        session.commit()

        service = DriverPlanningImportService(session)
        row_celny = ModulonPlanningRow(
            source_row=11, driver_number="", personnel_number="", first_name="Maciej",
            last_name="Celny", driver_group="", branch="", day=date(2026, 7, 1),
            source_status="FW", mapped_status="Freiwoche",
        )
        row_gawronski = ModulonPlanningRow(
            source_row=14, driver_number="", personnel_number="", first_name="Pawel",
            last_name="Gawronski", driver_group="", branch="", day=date(2026, 7, 1),
            source_status="FW", mapped_status="Freiwoche",
        )

        match, reason = service._find_driver(row_celny)
        assert match.id == celny.id
        assert reason == "Zweitname abweichend"
        match, reason = service._find_driver(row_gawronski)
        assert match.id == gawronski.id
        assert reason == "Name exakt/normalisiert"


def test_driver_merge_archives_source_and_transfers_absences():
    from datetime import datetime, timedelta
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.models.resource_absence import DriverAbsence
    from leipzigerflow.services.driver_service import DriverService

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as session:
        source = Driver(first_name="Maciej", last_name="Celny", city="Leipzig", personnel_number="17")
        target = Driver(first_name="Maciej Jan", last_name="Celny", city="Leipzig")
        session.add_all([source, target]); session.flush()
        absence = DriverAbsence(
            driver_id=source.id,
            starts_at=datetime(2026, 7, 1),
            ends_at=datetime(2026, 7, 2),
            reason="Freiwoche",
            active=True,
        )
        session.add(absence); session.commit()

        DriverService(session).merge(source, target)

        session.expire_all()
        assert session.get(Driver, source.id).active is False
        assert session.get(Driver, target.id).personnel_number == "17"
        stored = session.scalar(select(DriverAbsence))
        assert stored.driver_id == target.id


def test_modulon_matches_real_report_middle_name_and_single_typo():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningRow

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as session:
        celny = Driver(first_name="Maciej Jozef", last_name="Celny", city="Myszkow")
        schannak = Driver(first_name="Chris", last_name="Schannak", city="Malsch")
        session.add_all([celny, schannak]); session.commit()
        service = DriverPlanningImportService(session)

        row_celny = ModulonPlanningRow(
            source_row=11, driver_number="0103", personnel_number="", first_name="Maciej",
            last_name="Celny", driver_group="Fernverkehr", branch="Ettlingen",
            day=date(2026, 7, 1), source_status="FW", mapped_status="Freiwoche",
        )
        row_schannak = ModulonPlanningRow(
            source_row=40, driver_number="2312", personnel_number="", first_name="Chris",
            last_name="Schannnak", driver_group="Nahverkehr", branch="Ettlingen",
            day=date(2026, 7, 1), source_status="FW", mapped_status="Freiwoche",
        )

        match, reason = service._find_driver(row_celny)
        assert match.id == celny.id
        assert reason == "Zweitname abweichend"
        match, reason = service._find_driver(row_schannak)
        assert match.id == schannak.id
        assert reason == "leichte Schreibabweichung"


def test_modulon_persists_external_mapping_after_first_match():
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.models.external_mapping import ExternalMapping
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningPreview, ModulonPlanningRow

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as session:
        driver = Driver(first_name="Chris", last_name="Schannak", city="Malsch")
        session.add(driver); session.commit()
        row = ModulonPlanningRow(
            source_row=40, driver_number="2312", personnel_number="", first_name="Chris",
            last_name="Schannnak", driver_group="Nahverkehr", branch="Ettlingen",
            day=date(2026, 7, 1), source_status="FW", mapped_status="Freiwoche",
        )
        preview = ModulonPlanningPreview(month=date(2026, 7, 1), rows=[row], source_file="test.xlsx")
        service = DriverPlanningImportService(session)
        result = service.import_preview(preview)
        assert result.imported == 1
        mapping = session.scalar(select(ExternalMapping))
        assert mapping.external_id == "2312"
        assert mapping.internal_id == driver.id

        driver.last_name = "Komplett geändert"
        session.commit()
        match, reason = service._find_driver(row)
        assert match.id == driver.id
        assert reason == "gespeicherte Modulon-Zuordnung"



def test_modulon_repeated_status_rows_create_only_one_mapping_with_autoflush_disabled():
    from sqlalchemy import create_engine, func, select
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.models.external_mapping import ExternalMapping
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningPreview, ModulonPlanningRow

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False, autoflush=False) as session:
        driver = Driver(first_name="Roman Ryszard", last_name="Bialy", city="Ettlingen")
        session.add(driver)
        session.commit()
        rows = [
            ModulonPlanningRow(
                source_row=10, driver_number="0102", personnel_number="", first_name="Roman Ryszard",
                last_name="Bialy", driver_group="Fernverkehr", branch="Ettlingen",
                day=date(2026, 8, day), source_status="Urlaub", mapped_status="Urlaub",
            )
            for day in (1, 2, 3, 4)
        ]
        preview = ModulonPlanningPreview(month=date(2026, 8, 1), rows=rows, source_file="test.xlsx")

        result = DriverPlanningImportService(session).import_preview(preview)

        assert result.imported == 4
        assert result.mappings_created == 1
        assert result.mappings_updated == 0
        assert session.scalar(select(func.count()).select_from(ExternalMapping)) == 1


def test_modulon_existing_mapping_is_updated_not_inserted_again():
    from sqlalchemy import create_engine, func, select
    from sqlalchemy.orm import Session

    from leipzigerflow.database.base import Base
    import leipzigerflow.models  # noqa: F401
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.models.external_mapping import ExternalMapping
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningPreview, ModulonPlanningRow

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False, autoflush=False) as session:
        driver = Driver(first_name="Maciej Jozef", last_name="Celny", city="Myszkow")
        session.add(driver)
        session.flush()
        session.add(ExternalMapping(
            source_system="Modulon", entity_type="driver", external_id="0103",
            internal_id=driver.id, external_label="Alt", match_method="alt",
        ))
        session.commit()
        row = ModulonPlanningRow(
            source_row=11, driver_number="0103", personnel_number="", first_name="Maciej",
            last_name="Celny", driver_group="Fernverkehr", branch="Ettlingen",
            day=date(2026, 8, 10), source_status="Urlaub", mapped_status="Urlaub",
        )
        preview = ModulonPlanningPreview(month=date(2026, 8, 1), rows=[row], source_file="test.xlsx")

        result = DriverPlanningImportService(session).import_preview(preview)

        assert result.mappings_created == 0
        assert result.mappings_updated == 1
        assert session.scalar(select(func.count()).select_from(ExternalMapping)) == 1
        mapping = session.scalar(select(ExternalMapping))
        assert mapping.external_label == "Maciej Celny"


def test_modulon_contiguous_status_days_are_grouped_into_periods():
    from leipzigerflow.imports.modulon_resource_planner import ModulonPlanningRow
    from leipzigerflow.services.driver_planning_import_service import DriverPlanningImportService

    class FakeDriver:
        id = 7

    driver = FakeDriver()
    rows = [
        (
            ModulonPlanningRow(
                source_row=11,
                driver_number="0103",
                personnel_number="",
                first_name="Maciej",
                last_name="Celny",
                driver_group="Fernverkehr",
                branch="Ettlingen",
                day=date(2026, 8, day),
                source_status="Urlaub",
                mapped_status="Urlaub",
            ),
            driver,
            "manuelle Zuordnung",
        )
        for day in (10, 11, 12, 14)
    ]

    periods = DriverPlanningImportService._group_contiguous_rows(rows)

    assert len(periods) == 2
    assert [item[0].day.day for item in periods[0]] == [10, 11, 12]
    assert [item[0].day.day for item in periods[1]] == [14]
