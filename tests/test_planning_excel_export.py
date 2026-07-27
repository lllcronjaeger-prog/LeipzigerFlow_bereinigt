from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from enum import Enum
from types import SimpleNamespace

from openpyxl import load_workbook

from leipzigerflow.exports import export_dispatch_proposal, export_tours


class Mode(Enum):
    NEW = "Neue Tour bilden"


@dataclass
class Assignment:
    order_number: str = "A-100"
    mode: Mode = Mode.NEW
    source_tour_number: str = ""
    vehicle_label: str = "L-LL 100"
    driver_label: str = "Max Fahrer"
    loading_at: datetime = datetime(2026, 7, 22, 8, 0)
    available_again_at: datetime = datetime(2026, 7, 22, 15, 30)
    transfer_minutes: int = 20
    waiting_minutes: int = 10
    score: int = 88
    reasons: list[str] = field(default_factory=lambda: ["Geeigneter Aufbau", "Kurze Anfahrt"])
    alternatives: list = field(default_factory=list)
    confidence_label: str = "Hohe Sicherheit"
    confidence_percent: int = 90
    equivalent_best: bool = False


def test_dispatch_proposal_export_creates_expected_sheets(tmp_path):
    result = SimpleNamespace(
        created_at=datetime(2026, 7, 22, 7, 0),
        orders_total=1,
        assigned_count=1,
        extended_tour_count=0,
        new_tour_count=1,
        open_count=0,
        subcontractor_count=0,
        assignments=[Assignment()],
        unassigned=[],
    )
    target = tmp_path / "proposal.xlsx"
    export_dispatch_proposal(target, result, [])

    workbook = load_workbook(target, data_only=True)
    assert workbook.sheetnames == [
        "Planungsvorschlag",
        "Offene Aufträge",
        "Optimierungsvorschläge",
        "Freie Kapazitäten",
        "Alternativen",
        "Tourübersicht",
        "Tourdetails",
        "Kennzahlen",
        "Ressourcen",
    ]
    assert workbook["Planungsvorschlag"]["A9"].value == "A-100"
    assert workbook["Planungsvorschlag"]["D9"].value == "L-LL 100"


def test_tour_export_contains_overview_and_order_positions(tmp_path):
    location_a = SimpleNamespace(full_display="Werk Leipzig")
    location_b = SimpleNamespace(full_display="Kunde Berlin")
    customer = SimpleNamespace(display_name="Musterkunde GmbH")
    order = SimpleNamespace(
        order_number="A-200",
        customer_order_number="K-55",
        customer=customer,
        reference="Ref",
        status="Geplant",
        required_trailer_type="Plane, Koffer",
        loading_date=date(2026, 7, 22),
        loading_time_from=time(8, 0),
        loading_time_until=time(9, 0),
        loading_location=location_a,
        unloading_date=date(2026, 7, 22),
        unloading_time_from=time(12, 0),
        unloading_time_until=time(13, 0),
        unloading_location=location_b,
    )
    position = SimpleNamespace(position=1, transport_order=order)
    tour = SimpleNamespace(
        tour_number="T-2026-00001",
        tour_date=date(2026, 7, 22),
        planned_start_time=time(6, 0),
        status="Geplant",
        vehicle_display="L-LL 200",
        trailer_display="L-LL 900 / Plane",
        driver_display="Erika Fahrerin",
        remarks="Frühschicht",
        positions=[position],
    )
    target = tmp_path / "tours.xlsx"
    export_tours(target, [tour])

    workbook = load_workbook(target, data_only=True)
    assert workbook.sheetnames == ["Tourübersicht", "Tourpositionen"]
    assert workbook["Tourübersicht"]["A5"].value == "T-2026-00001"
    assert workbook["Tourpositionen"]["I5"].value == "A-200"
    assert workbook["Tourpositionen"]["R5"].value == "Werk Leipzig"
