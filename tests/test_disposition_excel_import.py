from datetime import date, time
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from leipzigerflow.database.base import Base
from leipzigerflow.imports.disposition_excel import build_preview, parse_address, parse_time_window
from leipzigerflow.models.customer import Customer
from leipzigerflow.models.driver import Driver
from leipzigerflow.models.location import Location
from leipzigerflow.models.tour import Tour
from leipzigerflow.models.tour_position import TourPosition
from leipzigerflow.models.transport_order import TransportOrder
from leipzigerflow.models.vehicle import Vehicle
from leipzigerflow.services.disposition_import_service import DispositionImportService


def make_file(path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Disposition"
    ws.append(["Ladetermin : 30.07.2026"])
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Ladetermin - Zeitfenster", "Liefertermin - Zeitfenster", "Frachtzahler", "Fahrer", "Stellplätze", "Paletten", "frachtpfl. Gewicht"])
    ws.append(["D1", "30.07.2026", "31.07.2026", "T100", "K100", "Lager A\nD 76149 Karlsruhe", "Kunde B\nD 68159 Mannheim", "KA-LL 8043", "08:00 - 09:00", "12:00 - 13:00", "Coca-Cola,BerlinKDFG", "Max Mustermann", 13.2, "EUR: 33", 22000])
    wb.save(path)


def test_parser_reads_disposition(tmp_path):
    path = tmp_path / "disp.xlsx"; make_file(path)
    preview = build_preview(path)
    assert len(preview.rows) == 1
    row = preview.rows[0]
    assert row.transport_number == "T100"
    assert row.loading_date == date(2026, 7, 30)
    assert row.loading_time_from == time(8, 0)
    assert row.pallets == 33
    assert row.has_planning


def test_address_and_time_parser():
    address = parse_address("LIT Lager\nD 76726 Germersheim")
    assert address.name == "LIT Lager"
    assert address.postal_code == "76726"
    assert parse_time_window("06:00 - 20:30") == (time(6, 0), time(20, 30))


def test_repeated_import_updates_without_duplicates(tmp_path):
    path = tmp_path / "disp.xlsx"; make_file(path)
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        service = DispositionImportService(session)
        first = service.import_rows(build_preview(path).rows)
        second = service.import_rows(build_preview(path).rows)
        assert first.orders_created == 1
        assert second.orders_updated == 1
        assert len(session.scalars(select(TransportOrder)).all()) == 1
        assert len(session.scalars(select(Location)).all()) == 2
        assert len(session.scalars(select(Customer)).all()) == 1
        assert len(session.scalars(select(Tour)).all()) == 1

from leipzigerflow.models.contractor import Contractor
from leipzigerflow.models.disposition_import_rule import DispositionImportRule


def test_repeated_headers_are_skipped(tmp_path):
    path = tmp_path / "headers.xlsx"
    wb = Workbook(); ws = wb.active
    headers = ["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Ladetermin - Zeitfenster", "Liefertermin - Zeitfenster", "Frachtzahler", "Fahrer", "Unternehmer"]
    ws.append(headers)
    ws.append(["D1", "30.07.2026", "30.07.2026", "T1", "", "A", "B", "", "", "", "K", "", ""])
    ws.append(headers)
    ws.append(["D2", "31.07.2026", "31.07.2026", "T2", "", "A", "B", "", "", "", "K", "", ""])
    wb.save(path)
    preview = build_preview(path)
    assert [row.transport_number for row in preview.rows] == ["T1", "T2"]


def test_default_storno_rule_does_not_import_order(tmp_path):
    path = tmp_path / "storno.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Beladeadresse", "Entladeadresse", "Frachtzahler", "Unternehmer"])
    ws.append(["D1", "30.07.2026", "30.07.2026", "T-STORNO", "A", "B", "K", "Storno laut Kunde"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        service = DispositionImportService(session)
        preview = service.mark_existing(build_preview(path))
        assert preview.rows[0].status == "Ignoriert"
        result = service.import_rows(preview.rows)
        assert result.ignored_by_rule == 1
        assert session.scalar(select(TransportOrder).where(TransportOrder.order_number == "T-STORNO")) is None


def test_open_disposition_rule_keeps_order_for_auto_dispatch_without_contractor(tmp_path):
    path = tmp_path / "open.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Beladeadresse", "Entladeadresse", "Frachtzahler", "Unternehmer"])
    ws.append(["D1", "30.07.2026", "30.07.2026", "T-OPEN", "A", "B", "K", "Luca"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(DispositionImportRule(name="Luca offen", field_name="Unternehmer", operator="ist gleich", comparison_value="Luca", action="Disposition offen", responsibility_hint="Luca", priority=5, active=True))
        session.commit()
        service = DispositionImportService(session)
        preview = service.mark_existing(build_preview(path))
        result = service.import_rows(preview.rows)
        order = session.scalar(select(TransportOrder).where(TransportOrder.order_number == "T-OPEN"))
        assert result.open_disposition_orders == 1
        assert order is not None
        assert order.contractor_id is None
        assert order.assignment_type == "Disposition offen"
        assert order.auto_dispatch_eligible is True
        assert order.planning_owner_hint == "Luca"
        assert session.scalars(select(Contractor)).all() == []


def test_customer_order_number_is_primary_sync_key_and_dossier_may_repeat(tmp_path):
    path = tmp_path / "multi.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Frachtzahler", "Fahrer", "Unternehmer", "Beladereferenz", "Entladereferenz"])
    ws.append(["D-4711", "31.07.2026", "31.07.2026", "T-SAMMEL", "KA-001", "A", "B", "KA-LL 8043", "K", "", "LLL", "L-1", "E-1"])
    ws.append(["D-4711", "31.07.2026", "31.07.2026", "T-SAMMEL", "KA-002", "A", "B", "KA-LL 8043", "K", "", "LLL", "L-2", "E-2"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine, autoflush=False) as session:
        service = DispositionImportService(session)
        result = service.import_rows(build_preview(path).rows)
        orders = session.scalars(select(TransportOrder).order_by(TransportOrder.customer_order_number)).all()
        positions = session.scalars(select(TourPosition).order_by(TourPosition.position)).all()
        assert result.orders_created == 2
        assert [o.customer_order_number for o in orders] == ["KA-001", "KA-002"]
        assert [o.dossier for o in orders] == ["D-4711", "D-4711"]
        assert [o.transport_number for o in orders] == ["T-SAMMEL", "T-SAMMEL"]
        assert [o.loading_reference for o in orders] == ["L-1", "L-2"]
        assert [p.position for p in positions] == [1, 2]


def test_reimport_updates_by_customer_order_number(tmp_path):
    path = tmp_path / "update_customer_order.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Frachtzahler", "Unternehmer"])
    ws.append(["D-1", "31.07.2026", "31.07.2026", "T-OLD", "KA-UNIQUE", "A", "B", "", "K", ""])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        service = DispositionImportService(session)
        first = service.import_rows(build_preview(path).rows)
        ws["D2"] = "T-NEW"; wb.save(path)
        second = service.import_rows(build_preview(path).rows)
        orders = session.scalars(select(TransportOrder)).all()
        assert first.orders_created == 1
        assert second.orders_updated == 1
        assert len(orders) == 1
        assert orders[0].transport_number == "T-NEW"


def test_storno_rule_is_rechecked_during_import_without_preview(tmp_path):
    path = tmp_path / "storno_direct.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Beladeadresse", "Entladeadresse", "Frachtzahler", "Unternehmer"])
    ws.append(["D1", "30.07.2026", "30.07.2026", "T-STORNO-DIRECT", "A", "B", "K", "  STORNO   LAUT KUNDE  "])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        result = DispositionImportService(session).import_rows(build_preview(path).rows)
        assert result.ignored_by_rule == 1
        assert session.scalars(select(TransportOrder)).all() == []


def test_storno_in_vehicle_column_removes_previously_imported_order(tmp_path):
    path = tmp_path / "storno_existing.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Frachtzahler", "Unternehmer"])
    ws.append(["D1", "31.07.2026", "31.07.2026", "T-1", "KA-STORNO", "A", "B", "", "K", "LLL"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine, autoflush=False) as session:
        service = DispositionImportService(session)
        first = service.import_rows(build_preview(path).rows)
        assert first.orders_created == 1
        ws["H2"] = "STORNO LAUT KUNDE"
        wb.save(path)
        second = service.import_rows(build_preview(path).rows)
        assert second.ignored_by_rule == 1
        assert session.scalars(select(TransportOrder)).all() == []
        assert session.scalars(select(TourPosition)).all() == []


def test_import_with_vehicle_assignment_does_not_call_routing_schedule(tmp_path, monkeypatch):
    from leipzigerflow.models.driver import Driver
    from leipzigerflow.models.vehicle import Vehicle
    from leipzigerflow.models.vehicle_resource_assignment import VehicleResourceAssignment
    from leipzigerflow.planner.time_planning import TimePlanningEngine

    path = tmp_path / "stammfahrer.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Frachtzahler", "Unternehmer"])
    ws.append(["D1", "31.07.2026", "31.07.2026", "T-2", "KA-2", "A", "B", "KA-LL 8043", "K", "LLL"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine, autoflush=False) as session:
        driver = Driver(first_name="Max", last_name="Stamm", active=True)
        vehicle = Vehicle(license_plate="KA-LL 8043", active=True)
        session.add_all([driver, vehicle]); session.flush()
        session.add(VehicleResourceAssignment(
            vehicle_id=vehicle.id, driver_id=driver.id, valid_from=date(2026, 7, 1), active=True
        ))
        session.commit()

        def fail_if_called(*_args, **_kwargs):
            raise AssertionError("Routing-/Zeitplanung darf im Import nicht aufgerufen werden")

        monkeypatch.setattr(TimePlanningEngine, "build_schedule", fail_if_called)
        result = DispositionImportService(session).import_rows(build_preview(path).rows)
        tour = session.scalar(select(Tour))
        assert result.orders_created == 1
        assert tour is not None
        assert tour.driver_id == driver.id
        assert len(tour.driver_assignments) == 1


def test_import_matches_driver_with_reversed_name_and_extra_spacing(tmp_path):
    path = tmp_path / "driver-match.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Fahrzeug", "Frachtzahler", "Fahrer", "Unternehmer"])
    ws.append(["D-1", "31.07.2026", "31.07.2026", "T-1", "KA-1", "A", "B", "KA-LL 9999", "K", "  Mustermann,   Max  ", "LLL"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        driver = Driver(first_name="Max", last_name="Mustermann", active=True)
        vehicle = Vehicle(license_plate="KA-LL 9999", active=True)
        session.add_all([driver, vehicle]); session.commit()
        service = DispositionImportService(session)
        service.import_rows(build_preview(path).rows)
        tour = session.scalar(select(Tour))
        assert tour is not None
        assert tour.driver_id == driver.id
        assert len(tour.driver_assignments) == 1
        assert tour.driver_assignments[0].driver_id == driver.id


def test_external_subcontractor_order_is_not_open_or_auto_dispatchable(tmp_path):
    path = tmp_path / "external.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Frachtzahler", "Unternehmer"])
    ws.append(["D-EXT", "31.07.2026", "31.07.2026", "T-EXT", "KA-EXT", "A", "B", "K", "Muster Transporte GmbH"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        result = DispositionImportService(session).import_rows(build_preview(path).rows)
        order = session.scalar(select(TransportOrder).where(TransportOrder.customer_order_number == "KA-EXT"))
        assert result.subcontractor_orders == 1
        assert order is not None
        assert order.assignment_type == "Subunternehmer"
        assert order.auto_dispatch_eligible is False
        assert order.status == "Extern vergeben"
        assert order.contractor is not None
        assert order.contractor.name == "Muster Transporte GmbH"
        assert session.scalars(select(TourPosition)).all() == []


def test_explicit_open_rule_can_reopen_named_subcontractor(tmp_path):
    path = tmp_path / "external_open.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Dossier", "Ladetermin", "Liefertermin", "Transportnummer", "Kundenauftragsnummer", "Beladeadresse", "Entladeadresse", "Frachtzahler", "Unternehmer"])
    ws.append(["D-OPEN", "31.07.2026", "31.07.2026", "T-OPEN-EXT", "KA-OPEN-EXT", "A", "B", "K", "Muster Transporte GmbH"])
    wb.save(path)
    engine = create_engine("sqlite+pysqlite:///:memory:"); Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(DispositionImportRule(
            name="Muster wieder offen", field_name="Unternehmer", operator="ist gleich",
            comparison_value="Muster Transporte GmbH", action="Disposition offen",
            priority=1, active=True,
        ))
        session.commit()
        DispositionImportService(session).import_rows(build_preview(path).rows)
        order = session.scalar(select(TransportOrder).where(TransportOrder.customer_order_number == "KA-OPEN-EXT"))
        assert order is not None
        assert order.assignment_type == "Disposition offen"
        assert order.auto_dispatch_eligible is True
        assert order.contractor_id is None
        assert order.status == "Neu"
