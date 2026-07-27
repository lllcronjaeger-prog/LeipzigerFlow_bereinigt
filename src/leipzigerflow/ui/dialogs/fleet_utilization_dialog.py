from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import (
    QDateEdit, QDialog, QFileDialog, QHBoxLayout, QLabel, QMessageBox,
    QPushButton, QTableWidget, QTableWidgetItem, QVBoxLayout,
)
from sqlalchemy.orm import Session

from leipzigerflow.services.fleet_utilization_service import FleetUtilizationService


class FleetUtilizationDialog(QDialog):
    """Separate Auswertung zu Fahrzeugauslastung und Tourmengen."""

    def __init__(self, session: Session, parent=None):
        super().__init__(parent)
        self.session = session
        self.service = FleetUtilizationService(session)
        self.snapshot = None
        self.setWindowTitle("Flottenauswertung · Auslastung und Touren")
        self.resize(1250, 760)

        root = QVBoxLayout(self)
        title = QLabel("<h2>Flottenauswertung</h2>")
        root.addWidget(title)

        filters = QHBoxLayout()
        filters.addWidget(QLabel("Von:"))
        self.date_from = QDateEdit(QDate.currentDate().addDays(-6))
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        filters.addWidget(self.date_from)
        filters.addWidget(QLabel("Bis:"))
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        filters.addWidget(self.date_to)
        refresh = QPushButton("Auswertung aktualisieren")
        refresh.clicked.connect(self.refresh)
        filters.addWidget(refresh)
        export = QPushButton("Auswertung als Excel")
        export.clicked.connect(self.export_excel)
        filters.addWidget(export)
        filters.addStretch()
        root.addLayout(filters)

        self.summary = QLabel()
        self.summary.setWordWrap(True)
        self.summary.setStyleSheet("font-size: 14px; font-weight: 600; padding: 8px;")
        root.addWidget(self.summary)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "Fahrzeug", "Art", "Klasse", "Touren", "Aufträge", "Geplant",
            "Kapazität", "Frei", "Auslastung", "Kapazitätshinweis",
        ])
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        root.addWidget(self.table, 1)

        footer = QHBoxLayout()
        footer.addStretch()
        close = QPushButton("Schließen")
        close.clicked.connect(self.accept)
        footer.addWidget(close)
        root.addLayout(footer)
        self.refresh()

    @staticmethod
    def _duration(minutes: int) -> str:
        return f"{minutes // 60}:{minutes % 60:02d} h"

    def refresh(self):
        self.snapshot = self.service.build(self.date_from.date().toPython(), self.date_to.date().toPython())
        s = self.snapshot
        self.summary.setText(
            f"Touren gesamt: {s.total_tours} · eigene Fahrzeuge: {s.own_tours} · "
            f"Fremdfahrzeuge: {s.foreign_tours} · durchschnittliche Auslastung: "
            f"{s.average_utilization:.1f} % · zusätzliche Tourkapazität: ca. {s.additional_tour_capacity}"
        )
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        for row_data in s.vehicles:
            row = self.table.rowCount()
            self.table.insertRow(row)
            values = [
                row_data.vehicle_label,
                row_data.ownership_type,
                row_data.vehicle_class,
                str(row_data.tour_count),
                str(row_data.order_count),
                self._duration(row_data.planned_minutes),
                self._duration(row_data.capacity_minutes),
                self._duration(row_data.free_minutes),
                f"{row_data.utilization_percent:.1f} %",
                row_data.capacity_hint,
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col in {3, 4, 5, 6, 7, 8}:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, col, item)
        self.table.resizeColumnsToContents()
        self.table.setSortingEnabled(True)

    def export_excel(self):
        if self.snapshot is None:
            self.refresh()
        default_name = f"Flottenauswertung_{self.snapshot.date_from:%Y-%m-%d}_bis_{self.snapshot.date_to:%Y-%m-%d}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Flottenauswertung exportieren", default_name, "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Flottenauswertung"
            headers = ["Fahrzeug", "Art", "Klasse", "Touren", "Aufträge", "Geplante Minuten", "Kapazität Minuten", "Freie Minuten", "Auslastung %", "Kapazitätshinweis"]
            ws.append(headers)
            for row in self.snapshot.vehicles:
                ws.append([row.vehicle_label, row.ownership_type, row.vehicle_class, row.tour_count, row.order_count, row.planned_minutes, row.capacity_minutes, row.free_minutes, round(row.utilization_percent, 1), row.capacity_hint])
            header_fill = PatternFill("solid", fgColor="1F4E78")
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            widths = [24, 16, 14, 10, 10, 18, 18, 16, 16, 48]
            for index, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + index)].width = width
            summary = wb.create_sheet("Kennzahlen")
            summary.append(["Kennzahl", "Wert"])
            summary.append(["Zeitraum", f"{self.snapshot.date_from:%d.%m.%Y} bis {self.snapshot.date_to:%d.%m.%Y}"])
            summary.append(["Touren gesamt", self.snapshot.total_tours])
            summary.append(["Touren eigene Fahrzeuge", self.snapshot.own_tours])
            summary.append(["Touren Fremdfahrzeuge", self.snapshot.foreign_tours])
            summary.append(["Aufträge eigene Fahrzeuge", self.snapshot.own_orders])
            summary.append(["Aufträge Fremdfahrzeuge", self.snapshot.foreign_orders])
            summary.append(["Durchschnittliche Auslastung %", round(self.snapshot.average_utilization, 1)])
            summary.append(["Zusätzliche Tourkapazität ca.", self.snapshot.additional_tour_capacity])
            for cell in summary[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
            summary.column_dimensions["A"].width = 34
            summary.column_dimensions["B"].width = 26
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(self, "Exportfehler", str(exc))
            return
        QMessageBox.information(self, "Export", f"Die Flottenauswertung wurde exportiert.\n\n{path}")
