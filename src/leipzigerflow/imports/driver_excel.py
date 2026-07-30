from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


_HEADER_ALIASES = {
    "matchcode": "match_code",
    "match code": "match_code",
    "match-code": "match_code",
    "anschrift": "address",
    "adresse": "address",
    "kontakt": "contact",
    "sonderberechtigungen": "special_permissions",
}
_COUNTRY_NAMES = {
    "D": "Deutschland",
    "DE": "Deutschland",
    "DEU": "Deutschland",
    "PL": "Polen",
    "POL": "Polen",
    "F": "Frankreich",
    "FR": "Frankreich",
    "FRA": "Frankreich",
    "CZ": "Tschechien",
    "CZE": "Tschechien",
    "SK": "Slowakei",
    "SVK": "Slowakei",
    "RO": "Rumänien",
    "ROU": "Rumänien",
    "HU": "Ungarn",
    "HUN": "Ungarn",
}


@dataclass(slots=True)
class DriverImportRow:
    source_row: int
    match_code: str = ""
    first_name: str = ""
    last_name: str = ""
    street: str = ""
    house_number: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = "Deutschland"
    phone: str = ""
    mobile: str = ""
    contact_raw: str = ""
    status: str = "Neu"
    errors: list[str] = field(default_factory=list)

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class DriverImportPreview:
    rows: list[DriverImportRow]
    sheet_name: str

    @property
    def valid_rows(self) -> list[DriverImportRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self) -> list[DriverImportRow]:
        return [row for row in self.rows if not row.is_valid]


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _normalize_header(value: object) -> str:
    text = re.sub(r"\s+", " ", _clean(value).lower())
    return _HEADER_ALIASES.get(text, text)


def _iter_xlsx(path: Path) -> tuple[str, list[dict[str, str]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(value) for value in next(rows, ())]
    result: list[dict[str, str]] = []
    for values in rows:
        if not any(_clean(value) for value in values):
            continue
        result.append({header: _clean(value) for header, value in zip(headers, values) if header})
    return sheet.title, result


def _iter_xls(path: Path) -> tuple[str, list[dict[str, str]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - depends on optional runtime package
        raise RuntimeError(
            "Für Dispoplan-Dateien im XLS-Format wird das Paket 'xlrd' benötigt."
        ) from exc
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    headers = [_normalize_header(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    result: list[dict[str, str]] = []
    for row_index in range(1, sheet.nrows):
        values = [sheet.cell_value(row_index, column) for column in range(sheet.ncols)]
        if not any(_clean(value) for value in values):
            continue
        result.append({header: _clean(value) for header, value in zip(headers, values) if header})
    return sheet.name, result


def _split_name(text: str) -> tuple[str, str]:
    parts = text.split()
    if len(parts) < 2:
        return (parts[0] if parts else "", "")
    return parts[0], " ".join(parts[1:])


def _split_street(text: str) -> tuple[str, str]:
    match = re.match(r"^(.*?)(?:\s+)(\d+[\w\-/]*)$", text.strip())
    if not match:
        return text.strip(), ""
    return match.group(1).strip(), match.group(2).strip()


def _parse_city_line(text: str) -> tuple[str, str, str]:
    value = re.sub(r"\s+", " ", text.strip())
    match = re.match(r"^(?:(?P<country>[A-Za-z]{1,3})\s+)?(?P<postal>[A-Za-z0-9-]{3,10})\s+(?P<city>.+)$", value)
    if not match:
        return "Deutschland", "", value
    code = (match.group("country") or "D").upper()
    return _COUNTRY_NAMES.get(code, code), match.group("postal"), match.group("city").strip()


def parse_address(value: str) -> dict[str, str]:
    lines = [re.sub(r"\s+", " ", line).strip() for line in value.split("\n") if line.strip()]
    name = lines[0] if lines else ""
    street_line = lines[1] if len(lines) > 1 else ""
    city_line = " ".join(lines[2:]) if len(lines) > 2 else ""
    first_name, last_name = _split_name(name)
    street, house_number = _split_street(street_line)
    country, postal_code, city = _parse_city_line(city_line)
    return {
        "first_name": first_name,
        "last_name": last_name,
        "street": street,
        "house_number": house_number,
        "postal_code": postal_code,
        "city": city,
        "country": country,
    }


def _phone_candidates(value: str) -> list[str]:
    candidates: list[str] = []
    for line in value.split("\n"):
        line = re.sub(r"(?i)\b(?:tel|telefon|mobil|handy)\s*[:.]?\s*", "", line).strip()
        if sum(character.isdigit() for character in line) < 6:
            continue
        normalized = re.sub(r"[^\d+]", " ", line)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        if normalized and normalized not in candidates:
            candidates.append(normalized)
    return candidates


def parse_contact(contact: str, special_permissions: str = "") -> tuple[str, str, str]:
    raw = contact.strip()
    candidates = _phone_candidates(raw)
    if not candidates and re.search(r"(?i)\btel(?:efon)?\b", special_permissions):
        candidates = _phone_candidates(special_permissions)
    phone = candidates[0] if candidates else ""
    mobile = candidates[1] if len(candidates) > 1 else ""
    return phone, mobile, raw


def build_preview(path: str | Path) -> DriverImportPreview:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".xlsx":
        sheet_name, source_rows = _iter_xlsx(source)
    elif suffix == ".xls":
        sheet_name, source_rows = _iter_xls(source)
    else:
        raise ValueError("Unterstützt werden Excel-Dateien im Format .xls und .xlsx.")

    rows: list[DriverImportRow] = []
    for source_row, raw in enumerate(source_rows, start=2):
        parsed = parse_address(raw.get("address", ""))
        phone, mobile, contact_raw = parse_contact(
            raw.get("contact", ""), raw.get("special_permissions", "")
        )
        row = DriverImportRow(
            source_row=source_row,
            match_code=raw.get("match_code", "").strip(),
            phone=phone,
            mobile=mobile,
            contact_raw=contact_raw,
            **parsed,
        )
        if not row.match_code:
            row.errors.append("MatchCode fehlt")
        if not row.first_name:
            row.errors.append("Vorname fehlt")
        if not row.last_name:
            row.errors.append("Nachname fehlt")
        rows.append(row)
    return DriverImportPreview(rows=rows, sheet_name=sheet_name)
