from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "dossier": "dossier",
    "ladetermin": "loading_date",
    "liefertermin": "unloading_date",
    "transportnummer": "transport_number",
    "kundenauftragsnummer": "customer_order_number",
    "beladeadresse": "loading_address",
    "entladeadresse": "unloading_address",
    "interne hinweise": "internal_notes",
    "fahrzeug": "vehicle",
    "beladereferenz": "loading_reference",
    "ladetermin - zeitfenster": "loading_window",
    "liefertermin - zeitfenster": "unloading_window",
    "entladereferenz": "unloading_reference",
    "hinweis": "notes",
    "stellplätze": "loading_meters",
    "stellplaetze": "loading_meters",
    "paletten": "pallets",
    "frachtpfl. gewicht": "weight_kg",
    "frachtzahler": "freight_payer",
    "fahrer": "driver",
    "unternehmer": "subcontractor",
}


@dataclass(slots=True)
class ParsedAddress:
    name: str = ""
    street: str = ""
    house_number: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = "Deutschland"

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.name.casefold(), self.postal_code.casefold(), self.city.casefold())


@dataclass(slots=True)
class DispositionImportRow:
    source_row: int
    dossier: str = ""
    transport_number: str = ""
    customer_order_number: str = ""
    loading_date: date | None = None
    unloading_date: date | None = None
    loading_time_from: time | None = None
    loading_time_until: time | None = None
    unloading_time_from: time | None = None
    unloading_time_until: time | None = None
    loading_address: ParsedAddress = field(default_factory=ParsedAddress)
    unloading_address: ParsedAddress = field(default_factory=ParsedAddress)
    freight_payer: str = ""
    vehicle: str = ""
    driver: str = ""
    loading_reference: str = ""
    unloading_reference: str = ""
    loading_meters: Decimal = Decimal("0")
    pallets: int = 0
    weight_kg: Decimal = Decimal("0")
    remarks: str = ""
    status: str = "Neu"
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    @property
    def is_cancelled(self) -> bool:
        text = f"{self.remarks} {self.vehicle}".casefold()
        return "storno" in text

    @property
    def has_planning(self) -> bool:
        vehicle = normalize_plate(self.vehicle)
        return bool(vehicle and vehicle not in {"-- KEINS --", "KEINS", "OFFEN"})


@dataclass(slots=True)
class DispositionImportPreview:
    rows: list[DispositionImportRow]
    sheet_name: str

    @property
    def valid_rows(self) -> list[DispositionImportRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self) -> list[DispositionImportRow]:
        return [row for row in self.rows if not row.is_valid]


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _header(value: object) -> str:
    text = re.sub(r"\s+", " ", _clean(value).casefold())
    return HEADER_ALIASES.get(text, text)


def _find_header_row(rows: list[list[object]]) -> int:
    for index, values in enumerate(rows[:25]):
        normalized = {_header(value) for value in values}
        if {"dossier", "loading_date", "transport_number", "loading_address"}.issubset(normalized):
            return index
    raise ValueError("Die Kopfzeile der Dispositionsliste wurde nicht gefunden.")


def _xlsx_rows(path: Path) -> tuple[str, list[dict[str, object]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    header_index = _find_header_row(raw_rows)
    headers = [_header(value) for value in raw_rows[header_index]]
    result: list[dict[str, object]] = []
    for values in raw_rows[header_index + 1:]:
        if not any(_clean(value) for value in values):
            continue
        row = {header: value for header, value in zip(headers, values) if header}
        if not _clean(row.get("transport_number")) and not _clean(row.get("dossier")):
            continue
        result.append(row)
    return sheet.title, result, header_index + 2


def _xls_rows(path: Path) -> tuple[str, list[dict[str, object]], int]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Für XLS-Dateien wird das Paket 'xlrd' benötigt.") from exc
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    raw_rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    header_index = _find_header_row(raw_rows)
    headers = [_header(value) for value in raw_rows[header_index]]
    result: list[dict[str, object]] = []
    for values in raw_rows[header_index + 1:]:
        if not any(_clean(value) for value in values):
            continue
        row = {header: value for header, value in zip(headers, values) if header}
        if not _clean(row.get("transport_number")) and not _clean(row.get("dossier")):
            continue
        result.append(row)
    return sheet.name, result, header_index + 2


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    for pattern in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def parse_time_window(value: object) -> tuple[time | None, time | None]:
    text = _clean(value)
    times = re.findall(r"(?<!\d)(\d{1,2}):(\d{2})(?!\d)", text)
    parsed = [time(int(hour), int(minute)) for hour, minute in times if int(hour) < 24 and int(minute) < 60]
    if not parsed:
        return None, None
    return parsed[0], parsed[1] if len(parsed) > 1 else parsed[0]


def parse_address(value: object) -> ParsedAddress:
    lines = [re.sub(r"\s+", " ", line).strip() for line in _clean(value).split("\n") if line.strip()]
    name = lines[0] if lines else ""
    address = ParsedAddress(name=name)
    for line in lines[1:]:
        match = re.match(r"^(?:(?P<country>[A-Za-z]{1,3})\s+)?(?P<postal>\d{4,6})\s+(?P<city>.+)$", line)
        if match:
            country = (match.group("country") or "D").upper()
            address.country = {"D": "Deutschland", "DE": "Deutschland", "F": "Frankreich", "FR": "Frankreich", "PL": "Polen"}.get(country, country)
            address.postal_code = match.group("postal")
            address.city = match.group("city").strip()
        elif not address.street:
            street_match = re.match(r"^(.*?)(?:\s+)(\d+[\w\-/]*)$", line)
            if street_match:
                address.street, address.house_number = street_match.group(1).strip(), street_match.group(2).strip()
            else:
                address.street = line
    return address


def _decimal(value: object) -> Decimal:
    text = _clean(value).replace(".", "").replace(",", ".") if isinstance(value, str) else _clean(value)
    try:
        return Decimal(text or "0")
    except InvalidOperation:
        return Decimal("0")


def _pallets(value: object) -> int:
    numbers = [int(number) for number in re.findall(r"\d+", _clean(value))]
    return sum(numbers)


def normalize_plate(value: str) -> str:
    text = re.sub(r"\s+", " ", value.upper().strip())
    return text.replace(" – ", "-").replace("–", "-")


def build_preview(path: str | Path) -> DispositionImportPreview:
    source = Path(path)
    if source.suffix.casefold() == ".xlsx":
        sheet_name, source_rows, first_row = _xlsx_rows(source)
    elif source.suffix.casefold() == ".xls":
        sheet_name, source_rows, first_row = _xls_rows(source)
    else:
        raise ValueError("Unterstützt werden Excel-Dateien im Format .xls und .xlsx.")

    rows: list[DispositionImportRow] = []
    for offset, raw in enumerate(source_rows):
        loading_from, loading_until = parse_time_window(raw.get("loading_window"))
        unloading_from, unloading_until = parse_time_window(raw.get("unloading_window"))
        row = DispositionImportRow(
            source_row=first_row + offset,
            dossier=_clean(raw.get("dossier")),
            transport_number=_clean(raw.get("transport_number")),
            customer_order_number=_clean(raw.get("customer_order_number")),
            loading_date=parse_date(raw.get("loading_date")),
            unloading_date=parse_date(raw.get("unloading_date")),
            loading_time_from=loading_from,
            loading_time_until=loading_until,
            unloading_time_from=unloading_from,
            unloading_time_until=unloading_until,
            loading_address=parse_address(raw.get("loading_address")),
            unloading_address=parse_address(raw.get("unloading_address")),
            freight_payer=_clean(raw.get("freight_payer")),
            vehicle=_clean(raw.get("vehicle")),
            driver=_clean(raw.get("driver")),
            loading_reference=_clean(raw.get("loading_reference")),
            unloading_reference=_clean(raw.get("unloading_reference")),
            loading_meters=_decimal(raw.get("loading_meters")),
            pallets=_pallets(raw.get("pallets")),
            weight_kg=_decimal(raw.get("weight_kg")),
            remarks="\n\n".join(filter(None, (_clean(raw.get("internal_notes")), _clean(raw.get("notes"))))),
        )
        if not row.transport_number:
            row.errors.append("Transportnummer fehlt")
        if not row.loading_date:
            row.errors.append("Ladetermin fehlt oder ist ungültig")
        if not row.unloading_date:
            row.errors.append("Liefertermin fehlt oder ist ungültig")
        if not row.loading_address.name:
            row.errors.append("Beladeadresse fehlt")
        if not row.unloading_address.name:
            row.errors.append("Entladeadresse fehlt")
        rows.append(row)
    return DispositionImportPreview(rows=rows, sheet_name=sheet_name)
