from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


_HEADER_ALIASES = {
    "kennzeichen kfz": "license_plate",
    "kennzeichen kfz.": "license_plate",
    "kennzeichen": "license_plate",
    "kfz kennzeichen": "license_plate",
    "kfz-kennzeichen": "license_plate",
}


@dataclass(slots=True)
class FleetImportRow:
    source_row: int
    license_plate: str = ""
    resource_type: str = "Trailer"
    match_code: str = ""
    status: str = "Neu"
    errors: list[str] = field(default_factory=list)

    @property
    def is_vehicle(self) -> bool:
        return self.resource_type == "Zugmaschine"

    @property
    def is_trailer(self) -> bool:
        return self.resource_type == "Trailer"

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class FleetImportPreview:
    rows: list[FleetImportRow]
    sheet_name: str

    @property
    def valid_rows(self) -> list[FleetImportRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self) -> list[FleetImportRow]:
        return [row for row in self.rows if not row.is_valid]

    @property
    def vehicle_rows(self) -> list[FleetImportRow]:
        return [row for row in self.rows if row.is_vehicle]

    @property
    def trailer_rows(self) -> list[FleetImportRow]:
        return [row for row in self.rows if row.is_trailer]


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_license_plate(value: object) -> str:
    text = _clean(value).upper()
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", " ", text)
    # Dispoplan liefert gelegentlich z. B. L-LL8068 ohne Leerzeichen vor der Nummer.
    text = re.sub(r"(?<=[A-Z])(?=\d)", " ", text)
    return text.strip()


def _normalize_header(value: object) -> str:
    text = re.sub(r"\s+", " ", _clean(value).lower())
    return _HEADER_ALIASES.get(text, text)


def _iter_xlsx(path: Path) -> tuple[str, list[tuple[int, str]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(value) for value in next(rows, ())]
    try:
        plate_index = headers.index("license_plate")
    except ValueError as exc:
        raise ValueError("Die Spalte 'Kennzeichen KfZ' wurde nicht gefunden.") from exc
    result: list[tuple[int, str]] = []
    for source_row, values in enumerate(rows, start=2):
        value = values[plate_index] if plate_index < len(values) else None
        if not _clean(value):
            continue
        result.append((source_row, _clean(value)))
    return sheet.title, result


def _iter_xls(path: Path) -> tuple[str, list[tuple[int, str]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - optional runtime package
        raise RuntimeError(
            "Für Dispoplan-Dateien im XLS-Format wird das Paket 'xlrd' benötigt."
        ) from exc
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    headers = [_normalize_header(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    try:
        plate_index = headers.index("license_plate")
    except ValueError as exc:
        raise ValueError("Die Spalte 'Kennzeichen KfZ' wurde nicht gefunden.") from exc
    result: list[tuple[int, str]] = []
    for row_index in range(1, sheet.nrows):
        value = sheet.cell_value(row_index, plate_index)
        if not _clean(value):
            continue
        result.append((row_index + 1, _clean(value)))
    return sheet.name, result


def is_vehicle_plate(license_plate: str) -> bool:
    compact = re.sub(r"[\s-]+", "", normalize_license_plate(license_plate))
    return compact.startswith("KALL")


def _number_part(license_plate: str) -> str:
    groups = re.findall(r"\d+", license_plate)
    return groups[-1] if groups else ""


def _registration_letters(license_plate: str) -> str:
    compact = normalize_license_plate(license_plate)
    match = re.match(r"^[A-ZÄÖÜ]{1,3}[\s-]+([A-ZÄÖÜ]{1,3})\s*\d+", compact)
    if match:
        return match.group(1)
    # Fallback für ungewöhnliche Schreibweisen: Buchstabengruppe direkt vor der Endnummer.
    match = re.search(r"([A-ZÄÖÜ]{1,3})\s*\d+\s*$", compact)
    return match.group(1) if match else ""


def _trailer_prefix(license_plate: str) -> str:
    letters = _registration_letters(license_plate)
    return letters[0] if letters else ""


def _looks_like_registration(license_plate: str) -> bool:
    return bool(re.match(r"^[A-ZÄÖÜ]{1,3}[\s-]+[A-ZÄÖÜ]{1,3}\s*\d+[A-Z]?\s*$", license_plate))


def assign_match_codes(rows: list[FleetImportRow]) -> None:
    """Vergibt kurze, eindeutige MatchCodes nach der vereinbarten Fuhrparkregel.

    Zugmaschinen erhalten die Endnummer. Trailer erhalten ebenfalls die Endnummer,
    solange diese nicht bereits von einer Zugmaschine oder einem anderen Trailer
    verwendet wird. Bei Überschneidungen wird dem Trailer der erste Buchstabe der
    mittleren Kennzeichengruppe vorangestellt (z. B. FR-H 4209 -> H4209).
    """
    vehicle_numbers = {
        _number_part(row.license_plate)
        for row in rows
        if row.is_vehicle and _number_part(row.license_plate)
    }
    trailer_number_counts: dict[str, int] = {}
    for row in rows:
        if row.is_trailer:
            number = _number_part(row.license_plate)
            if number:
                trailer_number_counts[number] = trailer_number_counts.get(number, 0) + 1

    used: dict[str, str] = {}
    for row in rows:
        number = _number_part(row.license_plate)
        if not number:
            continue
        if row.is_vehicle:
            candidate = number
        else:
            collision = number in vehicle_numbers or trailer_number_counts.get(number, 0) > 1
            candidate = f"{_trailer_prefix(row.license_plate)}{number}" if collision else number
        if not candidate:
            continue
        previous_plate = used.get(candidate.casefold())
        if previous_plate and previous_plate.casefold() != row.license_plate.casefold():
            row.errors.append(f"MatchCode {candidate} ist nicht eindeutig")
        else:
            used[candidate.casefold()] = row.license_plate
            row.match_code = candidate


def build_preview(path: str | Path) -> FleetImportPreview:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".xlsx":
        sheet_name, source_rows = _iter_xlsx(source)
    elif suffix == ".xls":
        sheet_name, source_rows = _iter_xls(source)
    else:
        raise ValueError("Unterstützt werden Excel-Dateien im Format .xls und .xlsx.")

    rows: list[FleetImportRow] = []
    seen_plates: set[str] = set()
    for source_row, raw_plate in source_rows:
        plate = normalize_license_plate(raw_plate)
        row = FleetImportRow(
            source_row=source_row,
            license_plate=plate,
            resource_type="Zugmaschine" if is_vehicle_plate(plate) else "Trailer",
        )
        if not _looks_like_registration(plate):
            row.errors.append("Kein gültiges Kfz-Kennzeichen")
        key = re.sub(r"[\s-]+", "", plate).casefold()
        if key in seen_plates:
            row.errors.append("Kennzeichen ist in der Datei doppelt vorhanden")
        seen_plates.add(key)
        rows.append(row)

    assign_match_codes(rows)
    for row in rows:
        if not row.match_code and not row.errors:
            row.errors.append("MatchCode konnte nicht ermittelt werden")
        if row.errors:
            row.status = "Fehler"
    return FleetImportPreview(rows=rows, sheet_name=sheet_name)
