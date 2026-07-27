from __future__ import annotations
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Ungültiges Datum: {value}")


def export_rows(path: str | Path, sheet: str, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook(); ws = wb.active; ws.title = sheet
    ws.append(headers)
    for row in rows: ws.append(row)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 40)
    wb.save(path)


def import_dicts(path: str | Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True); ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    result=[]
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in values): continue
        result.append(dict(zip(headers, values)))
    return result

parse_date = _date
