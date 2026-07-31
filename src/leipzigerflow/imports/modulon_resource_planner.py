from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable
import warnings

from openpyxl import load_workbook


KNOWN_STATUS_MAP = {
    "urlaub": "Urlaub",
    "fw": "Freiwoche",
    "freiwoche": "Freiwoche",
    "krank": "Krankheit",
    "krankheit": "Krankheit",
}


@dataclass(slots=True)
class ModulonPlanningRow:
    source_row: int
    driver_number: str
    personnel_number: str
    first_name: str
    last_name: str
    driver_group: str
    branch: str
    day: date
    source_status: str
    mapped_status: str
    errors: list[str] = field(default_factory=list)

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def is_valid(self) -> bool:
        return not self.errors and bool(self.day and self.source_status)


@dataclass(slots=True)
class ModulonPlanningPreview:
    month: date
    rows: list[ModulonPlanningRow]
    source_file: str

    @property
    def valid_rows(self) -> list[ModulonPlanningRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def unknown_statuses(self) -> set[str]:
        return {
            row.source_status
            for row in self.rows
            if row.source_status and row.source_status.casefold() not in KNOWN_STATUS_MAP
        }


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = _clean(value)
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date().replace(day=1) if fmt == "%m.%Y" else parsed.date()
        except ValueError:
            pass
    return None


def build_preview(path: str | Path) -> ModulonPlanningPreview:
    path = str(path)
    # Modulon exports may contain unsupported conditional-formatting extensions.
    # We only read cell values and never write the workbook back, so the openpyxl
    # warning is irrelevant for this import and can be hidden deliberately.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Conditional Formatting extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if "Zusatzinformationen" not in workbook.sheetnames:
            raise ValueError("Das Tabellenblatt 'Zusatzinformationen' wurde nicht gefunden.")
        sheet = workbook["Zusatzinformationen"]
        month = _excel_date(sheet["C2"].value)
        if month is None:
            raise ValueError("Der Berichtsmonat in Zelle C2 konnte nicht gelesen werden.")
        month = month.replace(day=1)

        date_columns: list[tuple[int, date]] = []
        for column in range(9, sheet.max_column + 1):
            day = _excel_date(sheet.cell(6, column).value) or _excel_date(sheet.cell(7, column).value)
            if day and day.year == month.year and day.month == month.month:
                date_columns.append((column, day))
        if not date_columns:
            raise ValueError("Im Modulon-Bericht wurden keine Tages-Spalten gefunden.")

        rows: list[ModulonPlanningRow] = []
        for row_index in range(8, sheet.max_row + 1):
            driver_number = _clean(sheet.cell(row_index, 1).value)
            personnel_number = _clean(sheet.cell(row_index, 2).value) or _clean(sheet.cell(row_index, 5).value)
            first_name = _clean(sheet.cell(row_index, 3).value)
            last_name = _clean(sheet.cell(row_index, 4).value)
            if not any((driver_number, personnel_number, first_name, last_name)):
                continue
            for column, day in date_columns:
                source_status = _clean(sheet.cell(row_index, column).value)
                if not source_status:
                    continue
                mapped = KNOWN_STATUS_MAP.get(source_status.casefold(), f"Modulon: {source_status}")
                errors: list[str] = []
                if not driver_number and not personnel_number and not (first_name and last_name):
                    errors.append("Keine eindeutige Fahrerkennung")
                rows.append(ModulonPlanningRow(
                    source_row=row_index,
                    driver_number=driver_number,
                    personnel_number=personnel_number,
                    first_name=first_name,
                    last_name=last_name,
                    driver_group=_clean(sheet.cell(row_index, 6).value),
                    branch=_clean(sheet.cell(row_index, 7).value),
                    day=day,
                    source_status=source_status,
                    mapped_status=mapped,
                    errors=errors,
                ))
        return ModulonPlanningPreview(month=month, rows=rows, source_file=path)
    finally:
        workbook.close()
