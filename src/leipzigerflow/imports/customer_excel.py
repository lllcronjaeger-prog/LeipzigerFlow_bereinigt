from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

_HEADER_ALIASES = {
    "name": "name",
    "matchcode": "match_code",
    "match code": "match_code",
    "anschrift": "address",
    "adresse": "address",
    "hauptkunde": "freight_payer",
    "frachtzahler": "freight_payer",
}
_COUNTRY_NAMES = {
    "D": "Deutschland", "DE": "Deutschland", "DEU": "Deutschland",
    "PL": "Polen", "F": "Frankreich", "FR": "Frankreich",
    "CZ": "Tschechien", "SK": "Slowakei", "A": "Österreich", "AT": "Österreich",
}


@dataclass(slots=True)
class CustomerImportRow:
    source_row: int
    name: str = ""
    match_code: str = ""
    street: str = ""
    house_number: str = ""
    postal_code: str = ""
    city: str = ""
    country: str = "Deutschland"
    freight_payer_match_code: str = ""
    freight_payer_name: str = ""
    status: str = "Neu"
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class CustomerImportPreview:
    rows: list[CustomerImportRow]
    sheet_name: str

    @property
    def valid_rows(self) -> list[CustomerImportRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self) -> list[CustomerImportRow]:
        return [row for row in self.rows if not row.is_valid]


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _normalize_header(value: object) -> str:
    text = re.sub(r"\s+", " ", _clean(value).lower())
    return _HEADER_ALIASES.get(text, text)


def _iter_xlsx(path: Path) -> tuple[str, list[dict[str, str]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(value) for value in next(rows, ())]
    result = []
    for values in rows:
        if not any(_clean(value) for value in values):
            continue
        result.append({header: _clean(value) for header, value in zip(headers, values) if header})
    return sheet.title, result


def _iter_xls(path: Path) -> tuple[str, list[dict[str, str]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Für XLS-Dateien wird das Paket 'xlrd' benötigt.") from exc
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    headers = [_normalize_header(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    result = []
    for row_index in range(1, sheet.nrows):
        values = [sheet.cell_value(row_index, column) for column in range(sheet.ncols)]
        if not any(_clean(value) for value in values):
            continue
        result.append({header: _clean(value) for header, value in zip(headers, values) if header})
    return sheet.name, result


def _split_street(text: str) -> tuple[str, str]:
    match = re.match(r"^(.*?)(?:\s+)(\d+[\w\-/]*)$", text.strip())
    return (match.group(1).strip(), match.group(2).strip()) if match else (text.strip(), "")


def _parse_city_line(text: str) -> tuple[str, str, str]:
    value = re.sub(r"\s+", " ", text.strip())
    match = re.match(r"^(?:(?P<country>[A-Za-z]{1,3})\s+)?(?P<postal>[A-Za-z0-9-]{3,10})\s+(?P<city>.+)$", value)
    if not match:
        return "Deutschland", "", value
    code = (match.group("country") or "D").upper()
    return _COUNTRY_NAMES.get(code, code), match.group("postal"), match.group("city").strip()


def parse_address(value: str) -> dict[str, str]:
    lines = [re.sub(r"\s+", " ", line).strip() for line in value.split("\n") if line.strip()]
    street_line = lines[-2] if len(lines) >= 2 else ""
    city_line = lines[-1] if lines else ""
    street, house_number = _split_street(street_line)
    country, postal_code, city = _parse_city_line(city_line)
    return {"street": street, "house_number": house_number, "postal_code": postal_code, "city": city, "country": country}


def parse_freight_payer(value: str) -> tuple[str, str]:
    text = value.strip()
    if not text:
        return "", ""
    if "|" in text:
        code, name = text.split("|", 1)
        return code.strip(), name.strip()
    return "", text


def build_preview(path: str | Path) -> CustomerImportPreview:
    source = Path(path)
    if source.suffix.lower() == ".xlsx":
        sheet_name, source_rows = _iter_xlsx(source)
    elif source.suffix.lower() == ".xls":
        sheet_name, source_rows = _iter_xls(source)
    else:
        raise ValueError("Unterstützt werden Excel-Dateien im Format .xls und .xlsx.")

    rows = []
    for source_row, raw in enumerate(source_rows, start=2):
        freight_code, freight_name = parse_freight_payer(raw.get("freight_payer", ""))
        row = CustomerImportRow(
            source_row=source_row,
            name=raw.get("name", "").strip(),
            match_code=raw.get("match_code", "").strip(),
            freight_payer_match_code=freight_code,
            freight_payer_name=freight_name,
            **parse_address(raw.get("address", "")),
        )
        if not row.name: row.errors.append("Name fehlt")
        if not row.match_code: row.errors.append("MatchCode fehlt")
        if not row.city: row.errors.append("Ort fehlt")
        rows.append(row)
    return CustomerImportPreview(rows=rows, sheet_name=sheet_name)
